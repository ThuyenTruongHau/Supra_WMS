#!/usr/bin/env python3
"""
Mock dữ liệu in Bacviet / Location, xuất PDF và/hoặc Excel cùng khổ A4.

Cài đặt (một lần):
    pip install playwright pillow
    playwright install chromium

Usage (từ thư mục backend/):
    python scripts/mock_print_export.py
    python scripts/mock_print_export.py --template location --format both
    python scripts/mock_print_export.py --template bacviet --quantity 12 --format xlsx
    python scripts/mock_print_export.py --format pdf --keep-html
"""

from __future__ import annotations

import argparse
import io
import json
import math
import sys
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from mock_print_layout import (
    BACVIET_LAYOUT,
    LOCATION_LAYOUT,
    mm_to_col_width,
    mm_to_row_height,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

BACKEND_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = BACKEND_ROOT / "app" / "static" / "templates"
DEFAULT_OUTPUT_DIR = BACKEND_ROOT / "scripts" / "output"

BACVIET_TEMPLATE = TEMPLATE_DIR / "template_bacviet.html"
LOCATION_TEMPLATE = TEMPLATE_DIR / "template_location.html"

QR_API = "https://api.qrserver.com/v1/create-qr-code/"


def build_bacviet_payload(
    quantity: int,
    *,
    sku: str,
    name: str,
    item_id: int,
) -> dict:
    date_part = datetime.now().strftime("%Y%m%d")
    display_code = f"{sku}-{date_part}"
    qr_ids = [f"{sku}-{date_part}-{index:08d}" for index in range(1, quantity + 1)]
    return {
        "mode": "preview",
        "__mock": True,
        "item_id": item_id,
        "quantity": quantity,
        "part_number": sku,
        "part_name": name,
        "qr_ids": qr_ids,
        "display_codes": [display_code] * quantity,
    }


def build_location_payload() -> dict:
    return {
        "__mock": True,
        "labels": [
            {
                "location_id": 1,
                "location_code": "R01_C01_L01_MC02_TP_PT",
                "location_name": "MC02_TP_PT",
                "qr_data": "R01_C01_L01_MC02_TP_PT",
            },
            {
                "location_id": 2,
                "location_code": "R02_C03_L02_MC05_KH_PT",
                "location_name": "MC05_KH_PT",
                "qr_data": "R02_C03_L02_MC05_KH_PT",
            },
            {
                "location_id": 3,
                "location_code": "R05_C10_L01_INBOUND-A",
                "location_name": "INBOUND-A",
                "qr_data": "R05_C10_L01_INBOUND-A",
            },
        ],
    }


def _vcc_logo_data_uri() -> str:
    logo_path = TEMPLATE_DIR / "logo_vcc.webp"
    if not logo_path.is_file():
        return ""
    import base64

    encoded = base64.b64encode(logo_path.read_bytes()).decode("ascii")
    return f"data:image/webp;base64,{encoded}"


def render_template_html(template_path: Path, payload: dict, data_key: str) -> str:
    if not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")

    template = template_path.read_text(encoding="utf-8")
    if template_path.name == "template_location.html":
        logo_uri = _vcc_logo_data_uri()
        if logo_uri:
            template = template.replace("__VCC_LOGO_DATA_URI__", logo_uri)
    script = (
        "<script>"
        f"window.{data_key} = {json.dumps(payload, ensure_ascii=False)};"
        "</script>\n"
    )
    marker = "  <script>\n    (function () {"
    if marker not in template:
        raise ValueError(f"Cannot inject print data into {template_path.name}")
    return template.replace(marker, f"{script}{marker}", 1)


def fetch_qr_png(data: str, *, size: int = 400) -> bytes:
    query = urllib.parse.urlencode({"size": f"{size}x{size}", "margin": "0", "data": data})
    url = f"{QR_API}?{query}"
    with urllib.request.urlopen(url, timeout=30) as response:
        return response.read()


def setup_a4_sheet(ws: Worksheet, *, margin_mm: float) -> None:
    margin_in = margin_mm / 25.4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = False
    ws.page_margins.left = margin_in
    ws.page_margins.right = margin_in
    ws.page_margins.top = margin_in
    ws.page_margins.bottom = margin_in
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = False


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _bacviet_labels_from_payload(payload: dict) -> list[dict]:
    quantity = int(payload["quantity"])
    part_number = str(payload["part_number"])
    part_name = str(payload["part_name"])
    qr_ids = list(payload["qr_ids"])
    display_codes = list(payload.get("display_codes") or qr_ids)
    return [
        {
            "part_number": part_number,
            "part_name": part_name,
            "qr_id": str(qr_ids[index]),
            "qr_display": str(display_codes[index] if index < len(display_codes) else qr_ids[index]),
        }
        for index in range(quantity)
    ]


def _require_pillow() -> None:
    try:
        import PIL  # noqa: F401
    except ImportError as exc:
        raise SystemExit(
            "Thiếu Pillow cho Excel QR. Cài đặt:\n  pip install pillow"
        ) from exc


def export_bacviet_xlsx(payload: dict, output_path: Path) -> None:
    _require_pillow()
    layout = BACVIET_LAYOUT
    labels = _bacviet_labels_from_payload(payload)
    wb = Workbook()
    wb.remove(wb.active)

    rows_per_label = 14
    cols_per_label = 4
    gap_cols = 1
    row_height_pt = mm_to_row_height(layout.label_height_mm / rows_per_label)
    col_width = mm_to_col_width(layout.label_width_mm / cols_per_label)

    thin = _thin_border()
    pages = math.ceil(len(labels) / layout.labels_per_page) if labels else 0

    for page_index in range(pages):
        ws = wb.create_sheet(title=f"Trang {page_index + 1}")
        setup_a4_sheet(ws, margin_mm=layout.page_margin_mm)

        for col_idx in range(layout.label_cols * (cols_per_label + gap_cols)):
            letter = get_column_letter(col_idx + 1)
            if gap_cols and col_idx % (cols_per_label + gap_cols) >= cols_per_label:
                ws.column_dimensions[letter].width = mm_to_col_width(layout.sheet_gap_mm)
            else:
                ws.column_dimensions[letter].width = col_width

        page_labels = labels[
            page_index * layout.labels_per_page : (page_index + 1) * layout.labels_per_page
        ]

        for slot, label in enumerate(page_labels):
            grid_col = slot % layout.label_cols
            grid_row = slot // layout.label_cols
            start_col = grid_col * (cols_per_label + gap_cols) + 1
            start_row = grid_row * rows_per_label + 1
            end_col = start_col + cols_per_label - 1
            end_row = start_row + rows_per_label - 1

            for row in range(start_row, end_row + 1):
                ws.row_dimensions[row].height = row_height_pt

            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col,
            )
            cell = ws.cell(start_row, start_col)
            cell.value = (
                "Phiếu sản phẩm / Product record sheet\n\n"
                f"Mã sản phẩm / Part number:\n{label['part_number']}\n\n"
                f"Tên sản phẩm / Part name:\n{label['part_name']}\n\n"
                f"QR: {label['qr_display']}"
            )
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Times New Roman", size=9)
            cell.border = thin

            qr_bytes = fetch_qr_png(label["qr_id"], size=180)
            qr_image = XLImage(io.BytesIO(qr_bytes))
            qr_image.width = 72
            qr_image.height = 72
            anchor_col = get_column_letter(end_col)
            ws.add_image(qr_image, f"{anchor_col}{max(start_row + 8, start_row)}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def export_location_xlsx(payload: dict, output_path: Path) -> None:
    _require_pillow()
    layout = LOCATION_LAYOUT
    labels = payload.get("labels") or []
    wb = Workbook()
    wb.remove(wb.active)

    title_rows = 3
    code_rows = 2
    qr_rows = 18
    total_rows = title_rows + code_rows + qr_rows
    title_height = mm_to_row_height(12)
    code_height = mm_to_row_height(8)
    qr_row_height = mm_to_row_height(
        (layout.content_height_mm - 12 - 8) / qr_rows
    )
    col_width = mm_to_col_width(layout.content_width_mm / 6)

    for index, label in enumerate(labels, start=1):
        ws = wb.create_sheet(title=f"{index}_{label['location_name'][:20]}")
        setup_a4_sheet(ws, margin_mm=layout.page_margin_mm)

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = col_width

        for row in range(1, title_rows + 1):
            ws.row_dimensions[row].height = title_height / title_rows
        for row in range(title_rows + 1, title_rows + code_rows + 1):
            ws.row_dimensions[row].height = code_height / code_rows
        for row in range(title_rows + code_rows + 1, total_rows + 1):
            ws.row_dimensions[row].height = qr_row_height

        ws.merge_cells(start_row=1, start_column=1, end_row=title_rows, end_column=6)
        title_cell = ws.cell(1, 1)
        title_cell.value = str(label["location_name"])
        title_cell.font = Font(name="Times New Roman", size=24, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells(
            start_row=title_rows + 1,
            start_column=1,
            end_row=title_rows + code_rows,
            end_column=6,
        )
        code_cell = ws.cell(title_rows + 1, 1)
        code_cell.value = str(label["location_code"])
        code_cell.font = Font(name="Times New Roman", size=13, bold=True)
        code_cell.alignment = Alignment(horizontal="center", vertical="center")

        qr_start_row = title_rows + code_rows + 1
        ws.merge_cells(
            start_row=qr_start_row,
            start_column=1,
            end_row=total_rows,
            end_column=6,
        )
        qr_cell = ws.cell(qr_start_row, 1)
        qr_cell.alignment = Alignment(horizontal="center", vertical="center")
        qr_cell.border = _thin_border()

        qr_data = str(label.get("qr_data") or label["location_code"])
        qr_bytes = fetch_qr_png(qr_data, size=800)
        qr_image = XLImage(io.BytesIO(qr_bytes))
        qr_image.width = 320
        qr_image.height = 320
        ws.add_image(qr_image, f"C{qr_start_row + 2}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def export_pdf(html: str, output_path: Path) -> Path:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise SystemExit(
            "Thiếu playwright cho PDF. Cài đặt:\n"
            "  pip install playwright\n"
            "  playwright install chromium"
        ) from exc

    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    html_path = output_path.with_suffix(".html")
    html_path.write_text(html, encoding="utf-8")

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch()
        page = browser.new_page()
        page.goto(html_path.as_uri(), wait_until="networkidle")
        page.emulate_media(media="print")
        page.pdf(
            path=str(output_path),
            format="A4",
            print_background=True,
            prefer_css_page_size=True,
            margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
        )
        browser.close()

    return html_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Mock in Bacviet/Location → PDF và/hoặc Excel A4.")
    parser.add_argument(
        "--template",
        choices=("bacviet", "location"),
        default="bacviet",
        help="Template cần mock (mặc định: bacviet)",
    )
    parser.add_argument(
        "--format",
        choices=("pdf", "xlsx", "both"),
        default="both",
        help="Định dạng output (mặc định: both)",
    )
    parser.add_argument("--quantity", type=int, default=9, help="Số phiếu Bacviet (mặc định: 9)")
    parser.add_argument("--sku", default="BGRD00074", help="Part number mock Bacviet")
    parser.add_argument(
        "--name",
        default="Thùng carton trà Bupnon TEA365 ô long vị đào (QR) 450ml",
        help="Tên sản phẩm mock Bacviet",
    )
    parser.add_argument("--item-id", type=int, default=1, help="Item id mock Bacviet")
    parser.add_argument("--output", type=Path, default=None, help="Đường dẫn file output (không gồm đuôi)")
    parser.add_argument("--keep-html", action="store_true", help="Giữ file HTML preview (PDF)")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.template == "bacviet" and args.quantity <= 0:
        raise SystemExit("--quantity phải lớn hơn 0")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_output = args.output or DEFAULT_OUTPUT_DIR / f"{args.template}_mock_{timestamp}"

    if args.template == "bacviet":
        payload = build_bacviet_payload(
            args.quantity,
            sku=args.sku,
            name=args.name,
            item_id=args.item_id,
        )
        html = render_template_html(BACVIET_TEMPLATE, payload, "__BACVIET_PRINT_DATA__")
        page_count = math.ceil(args.quantity / BACVIET_LAYOUT.labels_per_page)
    else:
        payload = build_location_payload()
        html = render_template_html(LOCATION_TEMPLATE, payload, "__LOCATION_PRINT_DATA__")
        page_count = math.ceil(len(payload["labels"]) / 2) if payload["labels"] else 0

    print(f"Template : {args.template}")
    print(f"Pages    : {page_count}")

    html_path: Path | None = None
    if args.format in {"pdf", "both"}:
        pdf_path = base_output.with_suffix(".pdf")
        html_path = export_pdf(html, pdf_path)
        print(f"PDF      : {pdf_path.resolve()}")
        if not args.keep_html and html_path is not None:
            html_path.unlink(missing_ok=True)
            html_path = None

    if args.format in {"xlsx", "both"}:
        xlsx_path = base_output.with_suffix(".xlsx")
        if args.template == "bacviet":
            export_bacviet_xlsx(payload, xlsx_path)
        else:
            export_location_xlsx(payload, xlsx_path)
        print(f"Excel    : {xlsx_path.resolve()}")

    if args.keep_html and html_path is not None:
        print(f"HTML     : {html_path.resolve()}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(130)
